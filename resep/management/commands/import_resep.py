import openpyxl
from django.core.management.base import BaseCommand
from resep.models import BarangJadi, MasterBahan, ResepBahanJadi

class Command(BaseCommand):
    help = 'Impor data dari file Excel'

    def add_arguments(self, parser):
        # Menambahkan argumen file_path pada perintah manage.py
        parser.add_argument('file_path', type=str, help='Path ke file Excel')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        
        # Memuat workbook
        workbook = openpyxl.load_workbook(file_path)
        
        # Mengasumsikan data ada di dua lembar pertama sesuai mekanisme ekspor
        worksheet_barang_jadi = workbook["Barang Jadi"]
        worksheet_resep = workbook["Resep"]

        # Memproses data untuk Barang Jadi
        for row in worksheet_barang_jadi.iter_rows(min_row=2, values_only=True):
            kode_barang_check = BarangJadi.objects.filter(kode_barang=row[1])
            if kode_barang_check:    
                print('Barang sudah ada, jadi di skip: ', kode_barang_check.first())
                barang_already_exist = True
            else:
                barang_jadi = BarangJadi.objects.create(
                    nama=row[0],
                    kode_barang=row[1],
                    harga_jual=row[2],
                    hpp=row[3],
                    master_roti=row[4]
                )
                barang_already_exist = False
            
        # Memproses data untuk Resep jika Barang Jadi belum ada
        if not barang_already_exist:
            for row in worksheet_resep.iter_rows(min_row=2, values_only=True):
                try:
                    barang_jadi = BarangJadi.objects.get(kode_barang=row[0])
                    master_bahan = MasterBahan.objects.get(kode_bahan=row[4])
                except BarangJadi.DoesNotExist:
                    print(f"BarangJadi dengan kode_barang {row[0]} Gagal Di simpan.")
                    continue
                except MasterBahan.DoesNotExist:
                    print(f"MasterBahan dengan kode_bahan {row[4]} tidak ditemukan.")
                    continue

                ResepBahanJadi.objects.create(
                    barang_jadi=barang_jadi,
                    master_bahan=master_bahan,
                    jumlah_pemakaian=row[2],
                )

        self.stdout.write(self.style.SUCCESS('Impor data selesai!'))
