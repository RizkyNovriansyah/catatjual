import csv
from django.core.management.base import BaseCommand
from resep.models import BarangJadi, ListPesanan, MasterBahan, Pesanan, Resep
# import pandas as pd
from django.core.management.base import BaseCommand
from openpyxl import Workbook
import os
from datetime import datetime

class Command(BaseCommand):
    help = 'Export data from database to an Excel file'

    def handle(self, *args, **kwargs):
        queryset_barang_jadi = BarangJadi.objects.filter(is_deleted=False)
        queryset_resep = Resep.objects.filter(is_deleted=False)
        workbook = Workbook()

        # Get active worksheet/tab
        ws1 = workbook.create_sheet("Barang Jadi")
        ws2 = workbook.create_sheet("Resep")
        
        colum_barang_jadi = [
            "nama",
            "kode_barang",
            "harga_jual",
            # "daftar_bahan",
            "hpp",
            "master_roti",
        ]
        
        colum_resep = [
            'barang_jadi_kode_barang',
            'barang_jadi_nama',
            'jumlah_pemakaian',
            'master_bahan_nama',
            'master_bahan_kode_bahan',
            'master_bahan_qty_keseluruhan',
            'master_bahan_qty_terkecil',
            'master_bahan_harga',
            'master_bahan_harga_kg',
            'master_bahan_harga_gram',
        ]
        
        row_num = 1
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(colum_barang_jadi, 1):
            cell = ws1.cell(row=row_num, column=col_num)
            cell.value = column_title
        
        row_num = 1
        for item in queryset_barang_jadi:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                item.nama,
                item.kode_barang,
                item.harga_jual,
                # item.daftar_bahan,
                item.hpp,
                item.master_roti,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = ws1.cell(row=row_num, column=col_num)
                cell.value = cell_value
                
        
        #------------Daftar resep----------------
        row_num = 1
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(colum_resep, 1):
            cell = ws2.cell(row=row_num, column=col_num)
            cell.value = column_title
            
            
        row_num = 1
        for item in queryset_resep:
            row_num += 1
            
            row = [
                item.barang_jadi.kode_barang,
                item.barang_jadi.nama,
                item.jumlah_pemakaian,
                item.master_bahan.nama,
                item.master_bahan.kode_bahan,
                item.master_bahan.qty_keseluruhan,
                item.master_bahan.qty_terkecil,
                item.master_bahan.harga,
                item.master_bahan.harga_kg,
                item.master_bahan.harga_gram,
            ]
            
            for col_num, cell_value in enumerate(row, 1):
                cell = ws2.cell(row=row_num, column=col_num)
                cell.value = cell_value
        
        current_date = datetime.now().strftime("%Y-%m-%d")
        file_path = os.path.join("data", f"export_resep_{current_date}.xlsx")

        workbook.save(filename=file_path)

        print(f"Export Resep done! File saved at: {file_path}")

        print("Export Resep done!")